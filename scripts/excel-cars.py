import os
import pandas as pd
from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup
from datetime import datetime

# File path to the Excel file
EXCEL_FILE_PATH = './data/cars.xlsx'

# URL and headers
URL = "https://tradingeconomics.com/canada/car-registrations"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9",
}

def fetch_recent_lei_data(url):
    """Fetch the 1st numeric value and the date from the 5th <td> element."""
    try:
        response = requests.get(url, headers=HEADERS)
        if response.status_code != 200:
            raise Exception(f"Request failed with status code {response.status_code}")

        soup = BeautifulSoup(response.content, "html.parser")

        # Find all <td> elements
        td_elements = soup.find_all("td")

        # Extract the 1st numeric value for the Average Home Price
        numeric_values = []
        for td in td_elements:
            text = td.get_text(strip=True)
            if text.replace('.', '', 1).isdigit():  # Check if the text is a valid number
                numeric_values.append(float(text))
            if len(numeric_values) == 1:  # Stop once we have the 9th numeric value
                break

        if len(numeric_values) < 9:
            raise Exception("Less than 9 numeric values found in <td> elements.")
        recent_value = numeric_values[-1]

        # Extract the date from the 5th <td> element
        if len(td_elements) < 5:
            raise Exception("Less than 5 <td> elements found.")
        recent_date_str = td_elements[4].get_text(strip=True)
        
        # Parse the date and reformat it as "YYYY-MM-DD"
        recent_date = datetime.strptime(recent_date_str, "%b %Y").replace(day=1).strftime("%Y-%m-%d")

        print(f"Fetched LEI data - Value: {recent_value}, Date: {recent_date}")
        return recent_date, recent_value
    except Exception as e:
        print(f"Error fetching LEI data: {e}")
        return None, None

def update_excel(file_path, recent_date, recent_value):
    """Update the Excel file with the most recent LEI data."""
    try:
        if not os.path.exists(file_path):
            print(f"File not found: {file_path}")
            return

        wb = load_workbook(file_path)
        sheet_name = "Data"
        if sheet_name not in wb.sheetnames:
            print(f"'{sheet_name}' sheet not found in the workbook.")
            return
        ws = wb[sheet_name]

        most_recent_date_in_excel = ws.cell(row=2, column=1).value
        if most_recent_date_in_excel:
            most_recent_date_in_excel = pd.to_datetime(most_recent_date_in_excel).strftime("%Y-%m-%d")
            most_recent_month = pd.to_datetime(most_recent_date_in_excel).month

        # Check if the most recent row needs to be replaced or if a new row should be added
        if most_recent_date_in_excel == recent_date:
            print(f"Recent date {recent_date} already exists in the Excel file. No update needed.")
            return
        elif most_recent_month != 12:
            # Replace row 2
            ws.cell(row=2, column=1, value=recent_date)  # Date
            ws.cell(row=2, column=2, value=recent_value)  # LEI Value
            print(f"Replaced row 2 with new data for date: {recent_date}")
        else:
            # Add a new row
            ws.insert_rows(2)
            ws.cell(row=2, column=1, value=recent_date)  # Date
            ws.cell(row=2, column=2, value=recent_value)  # LEI Value
            print(f"Added a new row for date: {recent_date}")

        # Update formulas in column C for all rows
        for row in range(2, ws.max_row):
            ws.cell(row=row, column=3, value=f"=(B{row}/B{row+1}-1)*100")

        # Save the workbook
        wb.save(file_path)
        print(f"Excel file updated successfully: {file_path}")
    except Exception as e:
        print(f"Error updating Excel file: {e}")

if __name__ == "__main__":
    recent_date, recent_value = fetch_recent_lei_data(URL)
    if recent_date and recent_value:
        update_excel(EXCEL_FILE_PATH, recent_date, recent_value)

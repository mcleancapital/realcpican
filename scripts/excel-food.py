import os
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from datetime import datetime

# URL of the page to scrape
url = "https://www150.statcan.gc.ca/n1/daily-quotidien/241220/t002a-eng.htm"

# Fetch the webpage content
response = requests.get(url)
soup = BeautifulSoup(response.content, 'html.parser')

# Parse the table to find the date and value
table = soup.find('table')
rows = table.find_all('tr')

# Extract the publication date from the table header
headers = rows[0].find_all('th')
raw_date = headers[3].text.strip()[:-1]  # Remove the last character from the date

# Add a space between the month and year if it's missing
if len(raw_date) > 4 and raw_date[-4:].isdigit():
    raw_date = raw_date[:-4] + " " + raw_date[-4:]

# Convert the raw date to "YYYY-MM-DD" format
formatted_date = datetime.strptime(raw_date, "%B %Y").strftime("%Y-%m-%d")

# Find the value for "Food and beverage retailers"
value = None
for row in rows:
    cells = row.find_all('td')
    if len(cells) > 0 and "Food and beverage retailers" in row.text:
        value = cells[2].text.strip().replace(',', '')  # Extract and clean the value
        break

# Ensure both date and value are extracted
if not value or not formatted_date:
    raise ValueError("Failed to extract required data.")

# File path for the Excel file
file_path = './data/food.xlsx'
sheet_name = "Data"  # Target sheet name

# Load existing data or create a new DataFrame if the file does not exist
if os.path.exists(file_path):
    existing_data = pd.read_excel(file_path, sheet_name=sheet_name)
else:
    existing_data = pd.DataFrame(columns=["Date", "Value", "% Change vs Last Year"])

# Ensure the "Date" column is in the same format as "formatted_date"
existing_data["Date"] = pd.to_datetime(existing_data["Date"]).dt.strftime("%Y-%m-%d")

# Check if the new date already exists
if formatted_date in existing_data["Date"].values:
    print(f"No new data added. Most recent date ({formatted_date}) already exists.")
    exit(0)

# Add the new data to the top of the DataFrame
new_row = pd.DataFrame({"Date": [formatted_date], "Value": [int(value)], "% Change vs Last Year": [None]})
updated_data = pd.concat([new_row, existing_data], ignore_index=True)

# Save the updated DataFrame to the Excel file
updated_data.to_excel(file_path, index=False, sheet_name=sheet_name)

# Load the workbook to calculate and insert static values dynamically
from openpyxl import load_workbook
wb = load_workbook(file_path)
ws = wb[sheet_name]

# Update column C with static values for the percentage change
for row in range(2, ws.max_row + 1):
    current_value = ws.cell(row=row, column=2).value  # Column B (current row value)
    offset_row = row + 12  # Row offset
    if offset_row <= ws.max_row:  # Ensure offset row exists
        offset_value = ws.cell(row=offset_row, column=2).value  # Column B (offset row value)
        if offset_value and offset_value != 0:
            percentage_change = (current_value / offset_value - 1) * 100
            ws.cell(row=row, column=3, value=round(percentage_change, 2))  # Static value in column C
        else:
            ws.cell(row=row, column=3, value=None)  # Handle division by zero or missing value
    else:
        ws.cell(row=row, column=3, value=None)  # No offset row available

# Save the workbook with static values in column C
wb.save(file_path)

print(f"Data for {formatted_date} and value {value} added to the top of the Excel file with static values in column C.")

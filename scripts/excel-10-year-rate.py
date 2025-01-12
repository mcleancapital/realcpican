import os
import pandas as pd
from yahoo_fin.stock_info import get_data
from openpyxl import load_workbook
from datetime import datetime

# File path to the Excel file
EXCEL_FILE_PATH = './data/10-year-treasury-rate.xlsx'

def fetch_recent_sp500_data():
    """Fetch the most recent S&P 500 data (date and price) from Yahoo Finance."""
    try:
        # Get historical data
        data = get_data("^TNX", start_date="2024-01-01")
        recent_date = data.index[-1].date()  # Most recent date
        recent_price = data['close'][-1]    # Most recent closing price
        print(f"Fetched data - Date: {recent_date}, Price: {recent_price}")
        return recent_date, recent_price
    except Exception as e:
        print(f"Error fetching recent S&P 500 data: {e}")
        return None, None

def update_excel(file_path, recent_date, recent_price):
    """Update the Excel file in the 'Data' tab and directly calculate values for column C."""
    try:
        # Check if file exists
        if not os.path.exists(file_path):
            print(f"File not found: {file_path}")
            return

        # Load the workbook and select the "Data" sheet
        wb = load_workbook(file_path)
        if "Data" not in wb.sheetnames:
            print(f"'Data' sheet not found in the workbook.")
            return
        ws = wb["Data"]

        # Check dates of the second and third rows
        second_row_date = ws.cell(row=2, column=1).value
        third_row_date = ws.cell(row=3, column=1).value

        if second_row_date:
            second_row_date = pd.to_datetime(second_row_date).date()
        if third_row_date:
            third_row_date = pd.to_datetime(third_row_date).date()

        # Determine whether to add a new row or update
        if second_row_date and third_row_date and second_row_date.month != third_row_date.month:
            # Add a new row
            ws.insert_rows(2)
            ws.cell(row=2, column=1, value=recent_date)
            ws.cell(row=2, column=2, value=recent_price)
            print(f"Inserted a new row with date {recent_date} and price {recent_price}.")
        else:
            # Update the second row
            ws.cell(row=2, column=1, value=recent_date)
            ws.cell(row=2, column=2, value=recent_price)
            print(f"Updated the second row with date {recent_date} and price {recent_price}.")

        # Directly calculate and update column C values
        max_row = ws.max_row
        for row in range(2, max_row + 1):
            reference_row = row + 12  # Reference row is 12 rows below
            if reference_row <= max_row:  # Only calculate if reference row exists
                current_value = ws.cell(row=row, column=2).value  # Get value from column B of the current row
                reference_value = ws.cell(row=reference_row, column=2).value  # Get value from column B of the reference row

                if current_value is not None and reference_value is not None and reference_value != 0:
                    calculated_value = ((current_value / reference_value) - 1) * 100  # Perform calculation
                    ws.cell(row=row, column=3, value=calculated_value)  # Set the calculated value
                else:
                    ws.cell(row=row, column=3, value=None)  # Clear cell if calculation cannot be performed
            else:
                ws.cell(row=row, column=3, value=None)  # Clear cell if reference row does not exist
        print(f"Updated values in column C for rows 2 to {max_row}.")

        # Save the workbook
        wb.save(file_path)
        print(f"Workbook saved successfully at {file_path}.")
    except Exception as e:
        print(f"Error updating Excel file: {e}")

if __name__ == "__main__":
    # Fetch the most recent S&P 500 data
    recent_date, recent_price = fetch_recent_sp500_data()

    # Update the Excel file in the "Data" tab
    if recent_date and recent_price:
        update_excel(EXCEL_FILE_PATH, recent_date, recent_price)

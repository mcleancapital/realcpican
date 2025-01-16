import requests
from openpyxl import Workbook, load_workbook
from datetime import datetime

# FRED API Setup
API_KEY = "dc0a72da11fb0ffc6f4dc8d31e7afbb5"  # Replace with your FRED API key
FRED_BASE = "https://api.stlouisfed.org/fred/series/observations"

# Function to fetch data from FRED API
def get_fred_data(series_id, api_key):
    params = {
        "series_id": series_id,
        "api_key": api_key,
        "file_type": "json",
    }
    response = requests.get(FRED_BASE, params=params)
    response.raise_for_status()
    data = response.json()
    if "observations" in data:
        return {obs["date"]: float(obs["value"]) for obs in data["observations"] if obs["value"] != "."}
    else:
        raise Exception(f"No data found for series {series_id}")

# Fetch High Yield Spread and 10-Year Treasury Yield
try:
    high_yield_spread = get_fred_data("BAMLH0A0HYM2EY", API_KEY)  # High Yield Spread
    treasury_yield = get_fred_data("DGS10", API_KEY)  # 10-Year Treasury Yield

    # Find the most recent common date
    common_dates = set(high_yield_spread.keys()).intersection(treasury_yield.keys())
    latest_date = max(common_dates)

    # Calculate the credit spread
    credit_spread = high_yield_spread[latest_date] - treasury_yield[latest_date]

    # Prepare Excel file
    excel_file = './data/credit-spread.xlsx'
    try:
        # Load existing workbook if it exists
        wb = load_workbook(excel_file)
        sheet = wb.active
    except FileNotFoundError:
        # Create a new workbook if it doesn't exist
        wb = Workbook()
        sheet = wb.active
        # Write headers
        sheet.append(["Date", "Value", "Percent Change"])

    # Check if the date already exists in the sheet
    existing_dates = [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1)]
    if latest_date not in existing_dates:
        # Add the new row
        sheet.insert_rows(2)
        sheet.cell(row=2, column=1, value=latest_date)
        sheet.cell(row=2, column=2, value=credit_spread)

        # Calculate the percentage change for the new row
        if sheet.max_row >= 251:
            try:
                reference_value_row = 252  # Row to use as reference for percentage change
                reference_value = sheet.cell(row=reference_value_row, column=2).value
                if reference_value:
                    percent_change = ((credit_spread / reference_value) - 1) * 100
                    sheet.cell(row=2, column=3, value=percent_change)
            except Exception as e:
                print(f"Error calculating percent change: {e}")

    # Save the workbook
    wb.save(excel_file)
    print(f"Credit spread ({credit_spread:.2f}) on {latest_date} saved to {excel_file}")

except Exception as e:
    print(f"Error: {e}")

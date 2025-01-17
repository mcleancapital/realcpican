import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
import requests
from bs4 import BeautifulSoup
from datetime import datetime

# File path to the Excel file
EXCEL_FILE_PATH = './data/wages.xlsx'

# URL and headers
URL = "https://ycharts.com/indicators/canada_compensation_per_hour_worked"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9",
}

def fetch_recent_cpi_data(url):
    """Fetch the most recent CPI data from the YCharts page."""
    try:
        response = requests.get(url, headers=HEADERS)
        if response.status_code != 200:
            raise Exception(f"Request failed with status code {response.status_code}")

        soup = BeautifulSoup(response.content, "html.parser")
        cpi_element = soup.find("div", class_="key-stat-title")
        if not cpi_element:
            raise Exception("Failed to find the element with class 'key-stat-title'.")

        cpi_text = cpi_element.get_text(strip=True)
        parts = cpi_text.split(maxsplit=2)
        if len(parts) < 2:
            raise Exception("Unexpected format for CPI data.")

        recent_value = float(parts[0])  # Extract TSI value
        date_text = parts[2].replace("for ", "")  # Clean up the date

        # Handle quarter-based dates
        if "Q" in date_text:
            quarter, year = date_text.split()
            year = int(year)
            quarter_end_month = {"Q1": "03-31", "Q2": "06-30", "Q3": "09-30", "Q4": "12-31"}
            recent_date = f"{year}-{quarter_end_month[quarter]}"
        else:
            raise Exception("Date format is not supported for conversion.")

        print(f"Fetched CPI data - Value: {recent_value}, Date: {recent_date}")
        return recent_date, recent_value
    except Exception as e:
        print(f"Error fetching CPI data: {e}")
        return None, None

def update_excel(file_path, recent_date, recent_value):
    """Update the Excel file with the most recent CPI data."""
    try:
        # Check if the file exists
        if not os.path.exists(file_path):
            print(f"File not found: {file_path}")
            return

        # Load the workbook
        wb = load_workbook(file_path)
        sheet_name = "Data"  # Assuming your data is in a sheet named 'Data'
        if sheet_name not in wb.sheetnames:
            print(f"'{sheet_name}' sheet not found in the workbook.")
            return
        ws = wb[sheet_name]

        # Check the most recent date in A2
        most_recent_date_in_excel = ws.cell(row=2, column=1).value
        if most_recent_date_in_excel:
            most_recent_date_in_excel = pd.to_datetime(most_recent_date_in_excel).strftime("%Y-%m-%d")

        # Add a new row only if the date is not already present
        if most_recent_date_in_excel == recent_date:
            print(f"Recent date {recent_date} already exists in the Excel file. No update needed.")
            return

        # Insert a new row
        ws.insert_rows(2)
        ws.cell(row=2, column=1, value=recent_date)  # Date
        ws.cell(row=2, column=2, value=recent_value)  # CPI Value

        # Update formulas in column C for all rows
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=3, value=f"=(B{row}/B{row+12}-1)*100")

        print(f"Added new data - Date: {recent_date}, Value: {recent_value}, Updated formulas in column C.")

        # Save the workbook
        wb.save(file_path)
        print(f"Excel file updated successfully: {file_path}")
    except Exception as e:
        print(f"Error updating Excel file: {e}")

if __name__ == "__main__":
    # Fetch the most recent CPI data
    recent_date, recent_value = fetch_recent_cpi_data(URL)

    # Update the Excel file if data is successfully fetched
    if recent_date and recent_value:
        update_excel(EXCEL_FILE_PATH, recent_date, recent_value)
